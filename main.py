import cv2
import face_recognition
import pickle
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import time

try:
    import mediapipe as mp
    MEDIAPIPE_OK = True
except Exception:
    MEDIAPIPE_OK = False

# =========================
# SETTINGS
# =========================
TEACHER_PASSWORD = "1234"
EXCEL_FILE = "Attendance.xlsx"
ENCODE_FILE = "EncodeFile.p"
STUDENTS_FILE = "students.csv"
IMAGES_FOLDER = "images"
CAMERA_INDEX = 0
FACE_MATCH_THRESHOLD = 0.50

# REQUIRED CHANGES
ATTENDANCE_WINDOW_MINUTES = 13
TEACHER_ACTIVATION_LIMIT_MINUTES = 15

FRAME_WIDTH = 960
FRAME_HEIGHT = 540
WINDOW_NAME = "Smart Attendance"

# =========================
# TIMETABLE
# =========================
TIMETABLE = {
    "Monday": [
        {"subject": "JAVA", "start": "08:20", "end": "09:00"},
        {"subject": "CC LAB", "start": "10:00", "end": "11:30"},
    ],
    "Tuesday": [
        {"subject": "CC", "start": "12:50", "end": "13:25"},
        {"subject": "JAVA", "start": "12:59", "end": "13:55"},
    ],
    "Wednesday": [
        {"subject": "PYTHON", "start": "08:00", "end": "09:00"},
        {"subject": "CC", "start": "10:00", "end": "11:00"},
    ],
    "Thursday": [
        {"subject": "DBMS", "start": "8:50", "end": "10:00"},
        {"subject": "PYTHON", "start": "18:53", "end": "19:59"},
    ],
    "Friday": [
        {"subject": "MATHS", "start": "08:00", "end": "09:00"},
        {"subject": "DBMS LAB", "start": "09:30", "end": "11:30"},
    ],
    "Saturday": [
        {"subject": "JAVA LAB", "start": "08:00", "end": "09:00"},
        {"subject": "PYTHON LAB", "start": "09:30", "end": "11:30"},
    ],
}

# =========================
# GLOBAL STATE
# =========================
attendance_active = False
window_closed = False
current_class = None
teacher_activated_at = ""
verified_students = set()
status_message = "Teacher not activated"
status_color = (0, 180, 255)
frame_count = 0
blink_done = False
last_photo_name = ""

# =========================
# MEDIAPIPE BLINK SETUP
# =========================
if MEDIAPIPE_OK:
    mp_face_mesh = mp.solutions.face_mesh
    face_mesh = mp_face_mesh.FaceMesh(
        max_num_faces=5,
        refine_landmarks=True,
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5,
    )
    LEFT_EYE = [33, 160, 158, 133, 153, 144]
    RIGHT_EYE = [362, 385, 387, 263, 373, 380]
else:
    face_mesh = None
    LEFT_EYE = []
    RIGHT_EYE = []

# =========================
# HELPERS
# =========================
def str_to_time(value: str):
    value = str(value).strip()
    if value == "24:00":
        value = "23:59"
    return datetime.strptime(value, "%H:%M").time()


def normalize_name(name: str) -> str:
    return str(name).strip().upper()


def sheet_name_safe(name: str) -> str:
    bad = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in bad:
        name = name.replace(ch, ' ')
    return name[:31]


def get_all_subjects():
    subjects = []
    for day_classes in TIMETABLE.values():
        for cls in day_classes:
            subject = str(cls["subject"]).strip().upper()
            if subject not in subjects:
                subjects.append(subject)
    return sorted(subjects)


def eye_ratio(landmarks, eye):
    p1 = landmarks[eye[0]]
    p2 = landmarks[eye[1]]
    p3 = landmarks[eye[2]]
    p4 = landmarks[eye[3]]
    p5 = landmarks[eye[4]]
    p6 = landmarks[eye[5]]

    vertical = abs(p2.y - p6.y) + abs(p3.y - p5.y)
    horizontal = abs(p1.x - p4.x)
    if horizontal == 0:
        return 1.0
    return vertical / (2.0 * horizontal)


def detect_blink(frame):
    if not MEDIAPIPE_OK:
        return True
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = face_mesh.process(rgb)
    if not results.multi_face_landmarks:
        return False
    for face_landmarks in results.multi_face_landmarks:
        landmarks = face_landmarks.landmark
        ear = (eye_ratio(landmarks, LEFT_EYE) + eye_ratio(landmarks, RIGHT_EYE)) / 2
        if ear < 0.24:
            return True
    return False


def build_class_payload(day, cls):
    now = datetime.now()
    start_dt = datetime.combine(now.date(), str_to_time(cls["start"]))
    end_dt = datetime.combine(now.date(), str_to_time(cls["end"]))
    activation_deadline = start_dt + timedelta(minutes=TEACHER_ACTIVATION_LIMIT_MINUTES)

    return {
        "day": day,
        "subject": cls["subject"].strip().upper(),
        "start_dt": start_dt,
        "end_dt": end_dt,
        "activation_deadline": activation_deadline,
        "attendance_end": datetime.now() + timedelta(minutes=ATTENDANCE_WINDOW_MINUTES),
        "class_time": f"{cls['start']} - {cls['end']}",
    }


def get_current_running_class():
    now = datetime.now()
    day = now.strftime("%A")
    if day not in TIMETABLE:
        return None
    for cls in TIMETABLE[day]:
        start_dt = datetime.combine(now.date(), str_to_time(cls["start"]))
        end_dt = datetime.combine(now.date(), str_to_time(cls["end"]))
        if start_dt <= now <= end_dt:
            return build_class_payload(day, cls)
    return None


def style_sheet(ws):
    headers = [
        "Name", "Roll No", "Date", "Day", "Subject", "Class Time",
        "Teacher Activated At", "Attendance Time", "Status"
    ]

    if ws.max_row == 1 and ws["A1"].value is None:
        ws.delete_rows(1, 1)

    if ws.max_row < 1:
        ws.append(headers)
    else:
        current_headers = [ws.cell(1, i).value for i in range(1, len(headers) + 1)]
        if current_headers != headers:
            if ws.max_row > 0:
                ws.delete_rows(1, ws.max_row)
            ws.append(headers)

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")

    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(1, idx)
        cell.value = title
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    widths = {
        "A": 18, "B": 12, "C": 14, "D": 14, "E": 18,
        "F": 18, "G": 22, "H": 18, "I": 18
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def ensure_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        default_ws = wb.active
        default_ws.title = "README"
        default_ws["A1"] = "Smart Attendance"
        default_ws["A2"] = "Subject-wise daily attendance + summary"
        wb.save(EXCEL_FILE)


def get_subject_sheet(wb, subject):
    sname = sheet_name_safe(subject)
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
    else:
        ws = wb[sname]
    style_sheet(ws)
    return ws


def get_summary_sheet(wb):
    sname = "SUMMARY"
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
    else:
        ws = wb[sname]

    subjects = get_all_subjects()

    headers = ["Name", "Roll No"]
    for subject in subjects:
        headers.extend([
            f"{subject} Present",
            f"{subject} Absent",
            f"{subject} %",
        ])
    headers.extend(["Total Present", "Total Absent", "Overall %"])

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    ws.append(headers)

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")

    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(1, idx)
        cell.value = title
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    for col in range(1, len(headers) + 1):
        if col == 1:
            ws.column_dimensions["A"].width = 18
        elif col == 2:
            ws.column_dimensions["B"].width = 12
        else:
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = 15

    return ws


def refresh_summary_sheet(student_dict):
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    summary_ws = get_summary_sheet(wb)
    subjects = get_all_subjects()

    student_summary = {}

    for name, roll in student_dict.items():
        uname = normalize_name(name)
        student_summary[uname] = {
            "roll": roll,
            "overall_present": 0,
            "overall_absent": 0,
            "subjects": {
                subject: {"present": 0, "absent": 0}
                for subject in subjects
            }
        }

    for sname in wb.sheetnames:
        if sname in ["README", "SUMMARY"]:
            continue

        ws = wb[sname]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            name = normalize_name(row[0])
            roll = row[1]
            subject = str(row[4]).strip().upper() if len(row) >= 5 and row[4] else sname.upper()
            status = str(row[8]).strip().upper() if len(row) >= 9 and row[8] is not None else ""

            if name not in student_summary:
                student_summary[name] = {
                    "roll": roll,
                    "overall_present": 0,
                    "overall_absent": 0,
                    "subjects": {
                        sub: {"present": 0, "absent": 0}
                        for sub in subjects
                    }
                }

            if subject not in student_summary[name]["subjects"]:
                student_summary[name]["subjects"][subject] = {"present": 0, "absent": 0}
                if subject not in subjects:
                    subjects.append(subject)

            if status == "PRESENT":
                student_summary[name]["overall_present"] += 1
                student_summary[name]["subjects"][subject]["present"] += 1
            elif status == "ABSENT":
                student_summary[name]["overall_absent"] += 1
                student_summary[name]["subjects"][subject]["absent"] += 1
            # CLASS NOT HELD ko percentage me include nahi karenge

    # alphabetical order by name
    for name in sorted(student_summary.keys()):
        row_data = [name, student_summary[name]["roll"]]

        for subject in subjects:
            present = student_summary[name]["subjects"].get(subject, {}).get("present", 0)
            absent = student_summary[name]["subjects"].get(subject, {}).get("absent", 0)
            total = present + absent
            percentage = 0.0 if total == 0 else (present / total) * 100

            row_data.extend([
                present,
                absent,
                f"{percentage:.2f}%"
            ])

        total_present = student_summary[name]["overall_present"]
        total_absent = student_summary[name]["overall_absent"]
        total_all = total_present + total_absent
        overall_percentage = 0.0 if total_all == 0 else (total_present / total_all) * 100

        row_data.extend([
            total_present,
            total_absent,
            f"{overall_percentage:.2f}%"
        ])

        summary_ws.append(row_data)

    wb.save(EXCEL_FILE)
    wb.close()


def row_matches_class(ws, row_num, student_name, date_str, subject, class_time):
    return (
        normalize_name(ws.cell(row_num, 1).value) == normalize_name(student_name)
        and str(ws.cell(row_num, 3).value) == date_str
        and str(ws.cell(row_num, 5).value).strip().upper() == str(subject).strip().upper()
        and str(ws.cell(row_num, 6).value) == class_time
    )


def ensure_student_row(ws, student_name, roll_no, date_str, day, subject, class_time, teacher_time, default_status="ABSENT"):
    student_name = normalize_name(student_name)
    subject = str(subject).strip().upper()

    for row in range(2, ws.max_row + 1):
        if row_matches_class(ws, row, student_name, date_str, subject, class_time):
            if not ws.cell(row, 2).value:
                ws.cell(row, 2).value = roll_no
            return row

    ws.append([
        student_name,
        roll_no,
        date_str,
        day,
        subject,
        class_time,
        teacher_time,
        "",
        default_status,
    ])
    return ws.max_row


def prefill_absent_rows(student_dict):
    if current_class is None:
        return

    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = get_subject_sheet(wb, current_class["subject"])
    date_str = datetime.now().strftime("%d-%m-%Y")

    for name, roll in student_dict.items():
        ensure_student_row(
            ws,
            name,
            roll,
            date_str,
            current_class["day"],
            current_class["subject"],
            current_class["class_time"],
            teacher_activated_at,
            default_status="ABSENT"
        )

    wb.save(EXCEL_FILE)
    wb.close()
    refresh_summary_sheet(student_dict)


def mark_class_not_held(student_dict, cls_payload):
    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = get_subject_sheet(wb, cls_payload["subject"])
    date_str = datetime.now().strftime("%d-%m-%Y")

    teacher_time_value = "NOT ACTIVATED IN 15 MIN"

    for name, roll in student_dict.items():
        row_num = ensure_student_row(
            ws,
            name,
            roll,
            date_str,
            cls_payload["day"],
            cls_payload["subject"],
            cls_payload["class_time"],
            teacher_time_value,
            default_status="CLASS NOT HELD"
        )
        ws.cell(row_num, 7).value = teacher_time_value
        ws.cell(row_num, 8).value = ""
        ws.cell(row_num, 9).value = "CLASS NOT HELD"

    wb.save(EXCEL_FILE)
    wb.close()
    refresh_summary_sheet(student_dict)


def mark_attendance(student_name, roll_no, student_dict):
    global verified_students, status_message, status_color

    if not attendance_active or window_closed or current_class is None:
        return False

    student_name = normalize_name(student_name)
    valid_students = [normalize_name(x) for x in student_dict.keys()]
    if student_name not in valid_students:
        return False

    unique_key = (
        student_name,
        datetime.now().strftime("%d-%m-%Y"),
        current_class["subject"],
        current_class["class_time"],
    )

    if unique_key in verified_students:
        status_message = f"Student already verified: {student_name}"
        status_color = (0, 255, 255)
        return False

    ensure_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = get_subject_sheet(wb, current_class["subject"])
    date_str = datetime.now().strftime("%d-%m-%Y")

    row_num = ensure_student_row(
        ws,
        student_name,
        roll_no,
        date_str,
        current_class["day"],
        current_class["subject"],
        current_class["class_time"],
        teacher_activated_at,
        default_status="ABSENT"
    )

    current_status = str(ws.cell(row_num, 9).value).strip().upper()

    if current_status == "CLASS NOT HELD":
        wb.close()
        status_message = "Class not held"
        status_color = (0, 0, 255)
        return False

    if current_status == "PRESENT" and ws.cell(row_num, 8).value:
        wb.close()
        verified_students.add(unique_key)
        status_message = f"Student already verified: {student_name}"
        status_color = (0, 255, 255)
        return False

    ws.cell(row_num, 7).value = teacher_activated_at
    if not ws.cell(row_num, 8).value:
        ws.cell(row_num, 8).value = datetime.now().strftime("%H:%M:%S")
    ws.cell(row_num, 9).value = "PRESENT"

    wb.save(EXCEL_FILE)
    wb.close()

    verified_students.add(unique_key)
    refresh_summary_sheet(student_dict)
    status_message = f"Attendance marked: {student_name}"
    status_color = (0, 255, 0)
    return True


def activate_teacher(student_dict):
    global attendance_active, window_closed, current_class, teacher_activated_at
    global status_message, status_color, verified_students

    current_class = get_current_running_class()
    if current_class is None:
        print("❌ Abhi koi running class nahi mili.")
        return False

    now = datetime.now()
    if now > current_class["activation_deadline"]:
        print("❌ Teacher ne 15 min ke andar class activate nahi ki.")
        print("✅ Sheet me CLASS NOT HELD mark kar diya gaya hai.")
        mark_class_not_held(student_dict, current_class)
        return False

    password = input("Enter Teacher Password: ").strip()
    if password != TEACHER_PASSWORD:
        print("❌ Wrong password")
        return False

    teacher_activated_at = datetime.now().strftime("%H:%M:%S")
    current_class["attendance_end"] = datetime.now() + timedelta(minutes=ATTENDANCE_WINDOW_MINUTES)
    attendance_active = True
    window_closed = False
    verified_students = set()
    status_message = "Teacher activated successfully"
    status_color = (0, 255, 0)

    prefill_absent_rows(student_dict)
    print(f"✅ Class: {current_class['subject']} | {current_class['class_time']}")
    print(f"✅ Activation deadline: {current_class['activation_deadline'].strftime('%H:%M:%S')}")
    print(f"✅ Attendance till: {current_class['attendance_end'].strftime('%H:%M:%S')}")
    return True


def load_student_photo(name_upper):
    base = name_upper.lower().strip()
    exts = [".jpg", ".jpeg", ".png", ".webp"]
    for ext in exts:
        path = os.path.join(IMAGES_FOLDER, base + ext)
        if os.path.exists(path):
            img = cv2.imread(path)
            if img is not None:
                return img
    return None


def open_camera():
    cap = cv2.VideoCapture(CAMERA_INDEX, cv2.CAP_DSHOW if os.name == 'nt' else cv2.CAP_ANY)
    if not cap.isOpened():
        raise RuntimeError("Camera open nahi hua")
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, FRAME_WIDTH)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, FRAME_HEIGHT)
    time.sleep(1)
    return cap


def draw_ui(frame, display_photo, recognized_info):
    h, w = frame.shape[:2]
    panel_h = 150
    panel = np.zeros((panel_h, w, 3), dtype=np.uint8)
    panel[:] = (12, 12, 12)

    title = "SMART ATTENDANCE"
    cv2.putText(frame, title, (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1.1, (255, 255, 255), 2)

    if current_class:
        cls = f"{current_class['subject']} | {current_class['class_time']}"
        cv2.putText(frame, cls, (20, 78), cv2.FONT_HERSHEY_SIMPLEX, 0.85, (0, 255, 255), 2)

        remaining = current_class['attendance_end'] - datetime.now()
        seconds = max(0, int(remaining.total_seconds()))
        mm = seconds // 60
        ss = seconds % 60
        win_text = "CLOSED" if window_closed else f"{mm:02d}:{ss:02d}"
        color = (0, 0, 255) if window_closed else (0, 255, 0)
        cv2.putText(frame, f"Window: {win_text}", (w - 260, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2)

    if window_closed:
        cv2.putText(frame, "ATTENDANCE WINDOW CLOSED", (w // 2 - 260, 90), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 3)

    left_x = 20
    mid_x = 340
    photo_x1 = w - 185
    photo_y1 = 15
    photo_x2 = w - 20
    photo_y2 = 135

    latest_name = recognized_info.get("name", "---")
    latest_roll = recognized_info.get("roll", "---")

    cv2.putText(panel, f"Name: {latest_name}", (left_x, 35), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)
    cv2.putText(panel, f"Roll: {latest_roll}", (left_x, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 0), 2)
    cls_text = current_class['subject'] if current_class else "No Active Class"
    cv2.putText(panel, f"Class: {cls_text}", (left_x, 105), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 255), 2)
    cv2.putText(panel, "Challenge: BLINK", (left_x, 138), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 200, 255), 2)

    cv2.putText(panel, f"Teacher Activated: {teacher_activated_at or '--:--:--'}", (mid_x, 35), cv2.FONT_HERSHEY_SIMPLEX, 0.85, (0, 255, 0), 2)
    till = current_class['attendance_end'].strftime('%H:%M:%S') if current_class else '--:--:--'
    cv2.putText(panel, f"Attendance Till: {till}", (mid_x, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.85, (0, 255, 0), 2)
    cv2.putText(panel, status_message[:40], (mid_x, 105), cv2.FONT_HERSHEY_SIMPLEX, 0.78, status_color, 2)
    cv2.putText(panel, "Q = Quit", (mid_x, 138), cv2.FONT_HERSHEY_SIMPLEX, 0.78, (180, 180, 180), 2)

    cv2.rectangle(panel, (photo_x1, photo_y1), (photo_x2, photo_y2), (100, 100, 100), 2)
    cv2.putText(panel, "PHOTO", (photo_x1 + 42, photo_y1 + 18), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (200, 200, 200), 1)

    if display_photo is not None:
        resized = cv2.resize(display_photo, (photo_x2 - photo_x1 - 4, photo_y2 - photo_y1 - 24))
        panel[photo_y1 + 22:photo_y2 - 2, photo_x1 + 2:photo_x2 - 2] = resized
    else:
        cv2.putText(panel, "No photo", (photo_x1 + 35, photo_y1 + 75), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (120, 120, 120), 2)

    return np.vstack((frame, panel))


# =========================
# LOAD DATA
# =========================
if not os.path.exists(STUDENTS_FILE):
    raise FileNotFoundError(f"{STUDENTS_FILE} file nahi mili")
if not os.path.exists(ENCODE_FILE):
    raise FileNotFoundError(f"{ENCODE_FILE} file nahi mili. Pehle encode.py run karo")

data = pd.read_csv(STUDENTS_FILE)
if "name" not in data.columns or "roll" not in data.columns:
    raise ValueError("students.csv me 'name' aur 'roll' columns hone chahiye")

data["name"] = data["name"].astype(str).str.strip().str.lower()
student_dict = dict(zip(data["name"], data["roll"]))
student_roll_upper = {normalize_name(k): v for k, v in student_dict.items()}

with open(ENCODE_FILE, "rb") as f:
    encode_list_known, class_names = pickle.load(f)
class_names = [os.path.splitext(x)[0].strip().lower() for x in class_names]

ensure_workbook()
refresh_summary_sheet(student_dict)

# =========================
# START
# =========================
print("Today:", datetime.now().strftime("%A %d-%m-%Y %H:%M:%S"))
print("Blink only challenge active")
print(f"Teacher activation limit: {TEACHER_ACTIVATION_LIMIT_MINUTES} minutes")
print(f"Student attendance window: {ATTENDANCE_WINDOW_MINUTES} minutes")

if not activate_teacher(student_dict):
    raise SystemExit

cap = open_camera()
cv2.namedWindow(WINDOW_NAME, cv2.WINDOW_NORMAL)
cv2.resizeWindow(WINDOW_NAME, FRAME_WIDTH, FRAME_HEIGHT + 150)

cached_faces = []
cached_names = []
cached_rolls = []
display_photo = None
recognized_info = {"name": "---", "roll": "---"}

while True:
    ok, frame = cap.read()
    if not ok or frame is None:
        continue

    frame = cv2.resize(frame, (FRAME_WIDTH, FRAME_HEIGHT))
    frame_count += 1

    if attendance_active and not window_closed and datetime.now() > current_class["attendance_end"]:
        attendance_active = False
        window_closed = True
        status_message = "Attendance window closed"
        status_color = (0, 0, 255)
        refresh_summary_sheet(student_dict)

    if not window_closed:
        blink_done = detect_blink(frame)

    process_this_frame = (frame_count % 3 == 0)
    if process_this_frame:
        small = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
        locations = face_recognition.face_locations(rgb_small, model="hog")
        encodings = face_recognition.face_encodings(rgb_small, locations)

        cached_faces = locations
        cached_names = []
        cached_rolls = []

        for encode_face in encodings:
            matches = face_recognition.compare_faces(
                encode_list_known,
                encode_face,
                tolerance=FACE_MATCH_THRESHOLD
            )
            face_dist = face_recognition.face_distance(encode_list_known, encode_face)

            name_upper = "UNKNOWN"
            roll_value = "---"

            if len(face_dist) > 0:
                best_idx = np.argmin(face_dist)
                if matches[best_idx]:
                    recognized_name = class_names[best_idx]
                    name_upper = normalize_name(recognized_name)
                    roll_value = student_roll_upper.get(name_upper, "---")

            cached_names.append(name_upper)
            cached_rolls.append(roll_value)

    current_detected_photo_name = ""
    for (top, right, bottom, left), name_upper, roll_value in zip(cached_faces, cached_names, cached_rolls):
        top, right, bottom, left = top * 4, right * 4, bottom * 4, left * 4
        color = (0, 255, 0) if name_upper != "UNKNOWN" else (0, 165, 255)
        cv2.rectangle(frame, (left, top), (right, bottom), color, 2)

        label_bg_bottom = max(28, top)
        cv2.rectangle(frame, (left, label_bg_bottom - 32), (right, label_bg_bottom), color, cv2.FILLED)
        cv2.putText(frame, name_upper, (left + 8, label_bg_bottom - 8), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)

        if name_upper != "UNKNOWN":
            recognized_info = {"name": name_upper, "roll": roll_value}
            current_detected_photo_name = name_upper
            unique_key = (
                name_upper,
                datetime.now().strftime("%d-%m-%Y"),
                current_class["subject"],
                current_class["class_time"],
            )

            if window_closed:
                msg = "CLOSED"
                msg_color = (0, 0, 255)
            elif unique_key in verified_students:
                msg = f"{name_upper} ALREADY VERIFIED"
                msg_color = (0, 255, 255)
                status_message = f"Student already verified: {name_upper}"
                status_color = (0, 255, 255)
            elif blink_done:
                if mark_attendance(name_upper, roll_value, student_dict):
                    msg = f"{name_upper} VERIFIED"
                else:
                    msg = f"{name_upper} VERIFIED"
                msg_color = (0, 255, 0)
            else:
                msg = "BLINK"
                msg_color = (0, 165, 255)

            cv2.putText(frame, msg, (left, bottom + 28), cv2.FONT_HERSHEY_SIMPLEX, 0.72, msg_color, 2)
        else:
            cv2.putText(frame, "UNKNOWN", (left, bottom + 28), cv2.FONT_HERSHEY_SIMPLEX, 0.72, (0, 165, 255), 2)

    if current_detected_photo_name:
        if last_photo_name != current_detected_photo_name:
            display_photo = load_student_photo(current_detected_photo_name)
            last_photo_name = current_detected_photo_name
    else:
        display_photo = None
        last_photo_name = ""
        recognized_info = {"name": "---", "roll": "---"}

    final = draw_ui(frame, display_photo, recognized_info)
    cv2.imshow(WINDOW_NAME, final)

    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()